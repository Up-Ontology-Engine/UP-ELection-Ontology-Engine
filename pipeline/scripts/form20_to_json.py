#!/usr/bin/env python3
"""
Convert all Excel files in 'data/raw/form20/' to JSON.
Output: data/processed/form20_json/{filename}.json

Handles three file types found in the directory:
  - Plain AC summary: 163.xls … 171.xls, 320.xls … 328.xls, AC320.xls … AC328.xls
  - Sub-part files:   320 (1).xls, 321 (2).xls, etc.  (Hindi-encoded content)
  - XLSX:             328.xlsx
"""

import json
import re
from pathlib import Path

import openpyxl
import xlrd

BASE_DIR = Path(__file__).resolve().parent.parent.parent
INPUT_DIR = BASE_DIR / "data" / "Form 20 Gorakhpur Data"
OUTPUT_DIR = BASE_DIR / "data" / "Form20_JSON"


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------


def _safe(v):
    """Normalise a cell value to a Python primitive."""
    if v is None:
        return None
    if isinstance(v, float):
        return int(v) if v == int(v) else v
    if isinstance(v, str):
        s = v.strip()
        return None if s in ("", "_", "-", "--") else s
    return v


def _read_xls(path: Path):
    """Return {sheet_name: [[cell, ...], ...]} for an .xls file."""
    wb = xlrd.open_workbook(str(path))
    result = {}
    for sh in wb.sheets():
        result[sh.name] = [[_safe(v) for v in sh.row_values(r)] for r in range(sh.nrows)]
    return result


def _read_xlsx(path: Path):
    """Return {sheet_name: [[cell, ...], ...]} for an .xlsx file."""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        result[name] = [[_safe(v) for v in row] for row in ws.iter_rows(values_only=True)]
    return result


def _find_col_num_row(rows):
    """Return index of the row that contains sequential column numbers (1,2,3…)."""
    for i, row in enumerate(rows):
        nums = [v for v in row if v is not None]
        if len(nums) >= 4 and nums[0] == 1 and nums[1] == 2 and nums[2] == 3:
            return i
    return None


def _is_total_row(row):
    first = next((v for v in row if v is not None), None)
    if isinstance(first, str) and "total" in first.lower():
        return True
    return False


def _header_text(header_rows):
    """Return a dict of metadata strings extracted from header rows."""
    meta = {}
    for row in header_rows:
        line = " ".join(str(v) for v in row if v is not None).strip()
        if not line:
            continue
        ll = line.lower()
        if "district" in ll:
            meta.setdefault("district", line)
        if "constituency" in ll or "assembly" in ll or "ac no" in ll or "vidhan sabha" in ll:
            meta.setdefault("constituency_info", line)
        if "election" in ll or "general" in ll or "nirvachan" in ll:
            meta.setdefault("election", line)
    return meta


# ---------------------------------------------------------------------------
# Candidate-column extraction
# ---------------------------------------------------------------------------


def _extract_candidates(header_rows, ncols):
    """
    Scan header rows for candidate columns.
    Returns (first_cand_col, candidates_list).
    candidates_list: [{serial, name, party}]
    """
    # Find the first column that holds candidate info by looking for
    # 'Tendered' or 'EPIC' in headers — candidates start after that.
    first_cand_col = None
    for row in header_rows:
        for j, v in enumerate(row):
            if v and isinstance(v, str):
                if "tendered" in v.lower():
                    first_cand_col = j + 1
                    break
        if first_cand_col is not None:
            break

    if first_cand_col is None:
        # Fallback: look for EPIC
        for row in header_rows:
            for j, v in enumerate(row):
                if v and isinstance(v, str) and "epic" in v.lower():
                    first_cand_col = j + 1
                    break
            if first_cand_col is not None:
                break

    if first_cand_col is None:
        first_cand_col = 9  # default

    # Find candidate name row (values after first_cand_col that are not 'Sl.'/'S.No.')
    name_row = None
    party_row = None
    for row in header_rows:
        vals_after = [v for v in row[first_cand_col:] if v is not None]
        if not vals_after:
            continue
        first_val = vals_after[0]
        if isinstance(first_val, str):
            if re.search(r"S\.?No\.?|Sl\.", first_val, re.I):
                party_row = row
            elif first_val not in ("Votes Secured", "Votes"):
                name_row = row

    candidates = []
    col = first_cand_col
    serial = 1
    while col + 2 <= ncols:
        name = None
        party = None
        if name_row and col < len(name_row):
            name = name_row[col]
        if party_row and col + 1 < len(party_row):
            party = party_row[col + 1]

        # Stop if both are None (beyond real candidate columns)
        if name is None and party is None and serial > 30:
            break

        candidates.append({"serial": serial, "name": name, "party": party, "_col": col})
        col += 3
        serial += 1

    return first_cand_col, candidates


# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------


def _parse_sheet(sheet_name, rows):
    if not rows:
        return None

    col_num_idx = _find_col_num_row(rows)
    if col_num_idx is None:
        return {
            "sheet": sheet_name,
            "note": "No column-number row found — likely Hindi-encoded or non-standard format",
            "row_count": len(rows),
            "sample": [r for r in rows[:6] if any(v is not None for v in r)],
        }

    header_rows = rows[:col_num_idx]
    data_rows = rows[col_num_idx + 1 :]
    ncols = max((len(r) for r in rows), default=0)

    metadata = _header_text(header_rows)
    first_cand_col, candidates = _extract_candidates(header_rows, ncols)

    # Detect "Other" gender turnout column by scanning header rows cols 5-8
    has_other = any(
        isinstance(row[j], str) and "other" in row[j].lower()
        for row in header_rows
        for j in range(5, min(9, len(row)))
    )

    polling_stations = []
    for row in data_rows:
        if all(v is None for v in row):
            continue
        if _is_total_row(row):
            continue

        def g(i):
            return row[i] if i < len(row) else None

        ac_no = g(0)
        ps_no = g(1)
        ps_name = g(2)

        # Skip non-data rows (e.g. sub-header repeats)
        if ac_no is None and ps_no is None:
            continue
        if isinstance(ps_no, str) and not ps_no.replace(".", "").isdigit():
            if ps_no.lower() in ("no.", "name"):
                continue

        if has_other:
            rec = {
                "ac_number": ac_no,
                "polling_station_number": ps_no,
                "polling_station_name": ps_name,
                "total_electors": g(3),
                "turnout_male": g(4),
                "turnout_female": g(5),
                "turnout_other": g(6),
                "turnout_total": g(7),
                "epic_voters": g(8),
                "tendered_votes": g(9),
            }
        else:
            rec = {
                "ac_number": ac_no,
                "polling_station_number": ps_no,
                "polling_station_name": ps_name,
                "total_electors": g(3),
                "turnout_male": g(4),
                "turnout_female": g(5),
                "turnout_total": g(6),
                "epic_voters": g(7),
                "tendered_votes": g(8) if first_cand_col > 8 else None,
            }

        # Candidate votes
        votes = []
        for cand in candidates:
            c = cand["_col"]
            raw_votes = g(c + 2)
            if isinstance(raw_votes, str):
                try:
                    raw_votes = int(float(raw_votes))
                except (ValueError, TypeError):
                    raw_votes = None

            entry = {"serial": cand["serial"], "votes": raw_votes}
            if cand["name"]:
                entry["candidate_name"] = cand["name"]
            if cand["party"]:
                entry["party"] = cand["party"]
            votes.append(entry)

        rec["candidate_votes"] = votes
        polling_stations.append(rec)

    return {
        "sheet": sheet_name,
        "metadata": metadata,
        "candidate_count": len(candidates),
        "polling_station_count": len(polling_stations),
        "polling_stations": polling_stations,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def convert(path: Path):
    if path.suffix.lower() == ".xlsx":
        sheets_data = _read_xlsx(path)
    else:
        sheets_data = _read_xls(path)

    parsed_sheets = []
    for name, rows in sheets_data.items():
        result = _parse_sheet(name, rows)
        if result:
            parsed_sheets.append(result)

    out = {
        "source_file": path.name,
        "sheets": parsed_sheets,
    }

    out_path = OUTPUT_DIR / (path.stem + ".json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    total_ps = sum(s.get("polling_station_count", 0) for s in parsed_sheets)
    return f"OK {path.name} → {total_ps} polling station records"


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(INPUT_DIR.glob("*.xls")) + sorted(INPUT_DIR.glob("*.xlsx"))
    # Skip the zip
    files = [f for f in files if f.suffix.lower() in (".xls", ".xlsx")]

    print(f"Found {len(files)} Excel files to convert")
    ok = fail = 0
    for f in files:
        try:
            msg = convert(f)
            print(f"  {msg}")
            ok += 1
        except Exception as e:
            print(f"  FAIL {f.name}: {e}")
            fail += 1

    print(f"\nDone. {ok} converted, {fail} failed.")


if __name__ == "__main__":
    main()
